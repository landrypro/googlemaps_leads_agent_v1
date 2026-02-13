from __future__ import annotations

from typing import Iterable
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..models import Lead

COLUMNS = [
    ("Name", "name"),
    ("Address", "address"),
    ("Phone", "phone"),
    ("Website", "website"),
    ("GoogleMapsUrl", "google_maps_url"),
    ("Latitude", "latitude"),
    ("Longitude", "longitude"),
    ("PlaceId", "place_id"),
]

def export_to_xlsx(leads: Iterable[Lead], out_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header
    ws.append([c[0] for c in COLUMNS])

    # Rows
    for lead in leads:
        # print(lead, 'my row')
        ws.append([getattr(lead, attr) for _, attr in COLUMNS])

    # Autosize columns (simple heuristic)
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(12, max_len + 2))

    wb.save(out_path)
    return out_path
